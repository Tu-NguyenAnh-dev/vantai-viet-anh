#!/usr/bin/env python3
"""
migrate_excel.py — Đọc file Excel thu chi xe và generate SQL INSERT

Hỗ trợ format: "Các xe TX.YYYY.xls(x)"
Sheet chính: "XE CHẠY TX" (X = số tháng)

Được gọi bởi import_excel.sh, không cần chạy trực tiếp.
Nếu muốn chạy thủ công:
  python3 migrate_excel.py <file.xlsx> --company-id <UUID> --out output.sql
"""

import sys
import argparse
import uuid
import re
from datetime import date
from typing import Optional

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: Thiếu openpyxl. Chạy: pip install openpyxl")


# ---------------------------------------------------------------------------
# Vehicle plate mapping (Xe XX → biển số)
# ---------------------------------------------------------------------------
def to_plate(raw: str) -> str:
    key = raw.strip().lower()
    m = re.match(r'^xe\s+(\w+)$', key)
    if m:
        return f"29C-{m.group(1).upper()}"
    return raw.strip()  # Kato 45, v.v.


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def sql_str(v) -> str:
    if v is None or str(v).strip() == "":
        return "NULL"
    return "'" + str(v).strip().replace("'", "''") + "'"


def num_val(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s.startswith("=") or s == "":
        return 0.0
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return 0.0


def normalize(name: str) -> str:
    return " ".join(str(name).split()).strip()


def detect_month_year(sheet_names: list[str]) -> tuple[int, int]:
    """Tự detect tháng/năm từ tên sheet, ví dụ 'XE CHẠY T7 ' → (7, 2025)."""
    for name in sheet_names:
        m = re.search(r'T(\d{1,2})', name, re.IGNORECASE)
        if m:
            month = int(m.group(1))
            # Tìm năm trong tên file hoặc sheet
            y = re.search(r'20(\d{2})', name)
            year = int(y.group(0)) if y else 2025
            return month, year
    return 7, 2025  # fallback


def find_main_sheet(wb) -> Optional[str]:
    """Tìm sheet chính (XE CHẠY TX)."""
    for name in wb.sheetnames:
        if re.search(r'XE\s+CH[AẠ]Y', name, re.IGNORECASE):
            return name
    return None


def to_date(day_val, month: int, year: int) -> Optional[str]:
    if day_val is None:
        return None
    try:
        day = int(float(str(day_val)))
        if 1 <= day <= 31:
            return date(year, month, day).isoformat()
    except (ValueError, TypeError, OverflowError):
        pass
    return None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", help="Đường dẫn file .xlsx")
    parser.add_argument("--company-id", required=True, help="UUID của company")
    parser.add_argument("--out", default="import_data.sql", help="File SQL output")
    args = parser.parse_args()

    company_id = args.company_id

    print(f"  Đọc file: {args.xlsx}")
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)

    # Tìm sheet chính
    sheet_name = find_main_sheet(wb)
    if not sheet_name:
        sys.exit(f"Không tìm thấy sheet 'XE CHẠY TX'. Các sheet có: {wb.sheetnames}")

    month, year = detect_month_year(wb.sheetnames)
    print(f"  Sheet: '{sheet_name}' | Tháng: {month}/{year}")

    ws = wb[sheet_name]

    # Parse dữ liệu (header row 3, data từ row 4)
    raw_rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        r = list(row[:17])
        if r[0] is not None and r[2] is not None:
            raw_rows.append(r)

    print(f"  Tìm thấy: {len(raw_rows)} dòng dữ liệu")

    # Collect unique entities
    vehicles: dict[str, str] = {}
    drivers: dict[str, str] = {}
    customers: dict[str, str] = {}
    DEFAULT_CUST = "Khách lẻ"

    for r in raw_rows:
        plate = to_plate(str(r[2]))
        vehicles.setdefault(plate, str(uuid.uuid4()))

        drv = normalize(str(r[1])) if r[1] else None
        if drv:
            drivers.setdefault(drv, str(uuid.uuid4()))

        cust = normalize(str(r[8])) if r[8] else DEFAULT_CUST
        customers.setdefault(cust or DEFAULT_CUST, str(uuid.uuid4()))

    customers.setdefault(DEFAULT_CUST, str(uuid.uuid4()))

    print(f"  Xe: {len(vehicles)} | Lái xe: {len(drivers)} | Khách: {len(customers)}")

    # Build SQL
    lines = []
    lines += [
        f"-- Import tháng {month}/{year} — company: {company_id}",
        "-- Tự động tạo bởi migrate_excel.py",
        "",
        "BEGIN;",
        "",
    ]

    # Vehicles
    lines.append("-- Vehicles")
    for plate, vid in vehicles.items():
        vtype = "cẩu bánh" if "kato" in plate.lower() else "cẩu"
        lines.append(
            f"INSERT INTO vehicles (id, company_id, license_plate, vehicle_type, status) "
            f"VALUES ('{vid}', '{company_id}', {sql_str(plate)}, {sql_str(vtype)}, 'active') "
            f"ON CONFLICT (company_id, license_plate) DO NOTHING;"
        )
    lines.append("")

    # Employees — dùng WHERE NOT EXISTS vì không có unique constraint trên full_name
    lines.append("-- Employees (lái xe)")
    for name, eid in drivers.items():
        lines.append(
            f"INSERT INTO employees (id, company_id, full_name, position, status) "
            f"SELECT '{eid}', '{company_id}', {sql_str(name)}, 'lái xe', 'active' "
            f"WHERE NOT EXISTS (SELECT 1 FROM employees WHERE company_id = '{company_id}' AND full_name = {sql_str(name)});"
        )
    lines.append("")

    # Customers — tương tự
    lines.append("-- Customers")
    for name, cid in customers.items():
        lines.append(
            f"INSERT INTO customers (id, company_id, name, status) "
            f"SELECT '{cid}', '{company_id}', {sql_str(name)}, 'active' "
            f"WHERE NOT EXISTS (SELECT 1 FROM customers WHERE company_id = '{company_id}' AND name = {sql_str(name)});"
        )
    lines.append("")

    # Trips (dùng DO block để lookup ID thực từ DB)
    lines += [
        "-- Trips",
        "DO $$",
        "DECLARE",
        "  v_id UUID; d_id UUID; c_id UUID; def_c UUID;",
        "BEGIN",
        f"  SELECT id INTO def_c FROM customers WHERE company_id = '{company_id}' AND name = {sql_str(DEFAULT_CUST)} LIMIT 1;",
        "",
    ]

    inserted = skipped = 0
    MULT = 1000  # nghìn đồng → VND

    for r in raw_rows:
        trip_date = to_date(r[0], month, year)
        if not trip_date:
            skipped += 1
            continue

        plate = to_plate(str(r[2]))
        drv = normalize(str(r[1])) if r[1] else None
        if not drv:
            skipped += 1
            continue

        cust = normalize(str(r[8])) if r[8] else DEFAULT_CUST

        # Tài chính
        rev     = num_val(r[4])  * MULT
        paid    = num_val(r[5])  * MULT
        toll    = num_val(r[9])  * MULT
        ticket  = num_val(r[10]) * MULT
        fine    = num_val(r[11]) * MULT
        repair  = num_val(r[13]) * MULT
        fuel    = num_val(r[14]) * MULT

        salary  = (rev - toll - ticket - fine) * 0.10
        other   = repair + fine + ticket
        profit  = rev - fuel - toll - salary - other

        # Notes
        mgr    = normalize(str(r[7])) if r[7] else None
        ghi_chu = normalize(str(r[15])) if len(r) > 15 and r[15] else None
        notes_parts = ([f"Quản lý: {mgr}"] if mgr else []) + ([ghi_chu] if ghi_chu else [])
        notes  = " | ".join(notes_parts) if notes_parts else None
        addr   = normalize(str(r[3])) if r[3] else None

        tid = str(uuid.uuid4())
        is_default_cust = (not cust or cust == DEFAULT_CUST)

        lines.append(f"  -- {trip_date} | {drv} | {plate}")
        lines.append(f"  SELECT id INTO v_id FROM vehicles  WHERE company_id = '{company_id}' AND license_plate = {sql_str(plate)} LIMIT 1;")
        lines.append(f"  SELECT id INTO d_id FROM employees WHERE company_id = '{company_id}' AND full_name = {sql_str(drv)} LIMIT 1;")
        if is_default_cust:
            lines.append(f"  c_id := def_c;")
        else:
            lines.append(f"  SELECT id INTO c_id FROM customers WHERE company_id = '{company_id}' AND name = {sql_str(cust)} LIMIT 1;")

        lines.append(
            f"  INSERT INTO trips "
            f"(id, company_id, trip_date, vehicle_id, driver_id, customer_id, "
            f"revenue, fuel_cost, toll_cost, driver_salary, other_costs, profit, "
            f"paid_amount, address, notes, status) VALUES ("
            f"'{tid}', '{company_id}', '{trip_date}', v_id, d_id, COALESCE(c_id, def_c), "
            f"{rev:.0f}, {fuel:.0f}, {toll:.0f}, {salary:.0f}, {other:.0f}, {profit:.0f}, "
            f"{paid:.0f}, {sql_str(addr)}, {sql_str(notes)}, 'completed') "
            f"ON CONFLICT DO NOTHING;"
        )
        lines.append("")
        inserted += 1

    lines += ["END $$;", "", "COMMIT;", ""]
    lines.append(f"-- Summary: {inserted} trips, {skipped} skipped | tháng {month}/{year}")

    with open(args.out, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"  SQL: {args.out}")
    print(f"  Kết quả: {inserted} trips ✓, {skipped} bỏ qua")


if __name__ == "__main__":
    main()
